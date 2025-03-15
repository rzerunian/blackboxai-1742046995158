import os
import json
from datetime import datetime
from openpyxl import Workbook, load_workbook

class ProjectManager:
    def __init__(self):
        """Initialize ProjectManager with necessary directories and files."""
        self.projects_dir = "projects"
        self.ensure_directory_exists()
        self.current_project = None

    def ensure_directory_exists(self):
        """Ensure the projects directory exists."""
        if not os.path.exists(self.projects_dir):
            os.makedirs(self.projects_dir)

    def create_project(self, name, description):
        """
        Create a new project with the given name and description.
        
        Args:
            name (str): Name of the project
            description (str): Description of the project
            
        Returns:
            bool: True if project was created successfully, False otherwise
        """
        try:
            # Validate project name
            if not name or not name.strip():
                raise ValueError("Nome do projeto não pode estar vazio")

            # Create project directory
            project_dir = os.path.join(self.projects_dir, name)
            if os.path.exists(project_dir):
                raise ValueError("Projeto já existe")
            
            os.makedirs(project_dir)

            # Create project metadata
            metadata = {
                "name": name,
                "description": description,
                "created_at": datetime.now().isoformat(),
                "last_modified": datetime.now().isoformat()
            }

            # Save metadata
            with open(os.path.join(project_dir, "metadata.json"), "w") as f:
                json.dump(metadata, f, indent=4)

            # Create initial Excel files
            self._create_initial_files(project_dir)

            return True

        except Exception as e:
            print(f"Erro ao criar projeto: {str(e)}")
            return False

    def _create_initial_files(self, project_dir):
        """
        Create initial Excel files for the project.
        
        Args:
            project_dir (str): Path to the project directory
        """
        # Create CapEx file
        wb_capex = Workbook()
        ws = wb_capex.active
        ws.title = "CapEx"
        headers = ["TAG", "Descrição", "Quantidade", "Valor Unitário", "Valor Total"]
        ws.append(headers)
        wb_capex.save(os.path.join(project_dir, "capex.xlsx"))

        # Create OpEx file
        wb_opex = Workbook()
        ws = wb_opex.active
        ws.title = "OpEx"
        ws.append(headers)
        wb_opex.save(os.path.join(project_dir, "opex.xlsx"))

        # Create Receitas file
        wb_receitas = Workbook()
        ws = wb_receitas.active
        ws.title = "Receitas"
        ws.append(headers)
        wb_receitas.save(os.path.join(project_dir, "receitas.xlsx"))

        # Create Config file
        wb_config = Workbook()
        ws = wb_config.active
        ws.title = "Configuração"
        ws.append(["Parâmetro", "Valor"])
        ws.append(["TMA", "0.0"])
        ws.append(["IR", "0.0"])
        ws.append(["CSLL", "0.0"])
        wb_config.save(os.path.join(project_dir, "config.xlsx"))

    def list_projects(self):
        """
        List all available projects.
        
        Returns:
            list: List of project names
        """
        try:
            projects = []
            for item in os.listdir(self.projects_dir):
                if os.path.isdir(os.path.join(self.projects_dir, item)):
                    metadata_path = os.path.join(self.projects_dir, item, "metadata.json")
                    if os.path.exists(metadata_path):
                        with open(metadata_path, "r") as f:
                            metadata = json.load(f)
                            projects.append(metadata["name"])
            return projects
        except Exception as e:
            print(f"Erro ao listar projetos: {str(e)}")
            return []

    def load_project(self, project_name):
        """
        Load a project and its associated files.
        
        Args:
            project_name (str): Name of the project to load
            
        Returns:
            dict: Project data including metadata and file paths
        """
        try:
            project_dir = os.path.join(self.projects_dir, project_name)
            if not os.path.exists(project_dir):
                raise ValueError("Projeto não encontrado")

            # Load metadata
            with open(os.path.join(project_dir, "metadata.json"), "r") as f:
                metadata = json.load(f)

            # Create project data structure
            project_data = {
                "metadata": metadata,
                "files": {
                    "capex": os.path.join(project_dir, "capex.xlsx"),
                    "opex": os.path.join(project_dir, "opex.xlsx"),
                    "receitas": os.path.join(project_dir, "receitas.xlsx"),
                    "config": os.path.join(project_dir, "config.xlsx")
                }
            }

            self.current_project = project_data
            return project_data

        except Exception as e:
            print(f"Erro ao carregar projeto: {str(e)}")
            return None

    def save_project(self, project_name, data):
        """
        Save project data to files.
        
        Args:
            project_name (str): Name of the project
            data (dict): Project data to save
            
        Returns:
            bool: True if saved successfully, False otherwise
        """
        try:
            project_dir = os.path.join(self.projects_dir, project_name)
            if not os.path.exists(project_dir):
                raise ValueError("Projeto não encontrado")

            # Update metadata
            metadata_path = os.path.join(project_dir, "metadata.json")
            with open(metadata_path, "r") as f:
                metadata = json.load(f)
            
            metadata["last_modified"] = datetime.now().isoformat()
            
            with open(metadata_path, "w") as f:
                json.dump(metadata, f, indent=4)

            # Save data to respective Excel files
            if "capex" in data:
                self._save_excel_data(os.path.join(project_dir, "capex.xlsx"), data["capex"])
            if "opex" in data:
                self._save_excel_data(os.path.join(project_dir, "opex.xlsx"), data["opex"])
            if "receitas" in data:
                self._save_excel_data(os.path.join(project_dir, "receitas.xlsx"), data["receitas"])
            if "config" in data:
                self._save_excel_data(os.path.join(project_dir, "config.xlsx"), data["config"])

            return True

        except Exception as e:
            print(f"Erro ao salvar projeto: {str(e)}")
            return False

    def _save_excel_data(self, file_path, data):
        """
        Save data to an Excel file.
        
        Args:
            file_path (str): Path to the Excel file
            data (list): Data to save
        """
        wb = Workbook()
        ws = wb.active
        for row in data:
            ws.append(row)
        wb.save(file_path)

    def get_current_project(self):
        """
        Get the currently loaded project.
        
        Returns:
            dict: Current project data or None if no project is loaded
        """
        return self.current_project
