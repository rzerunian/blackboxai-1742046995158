from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

class ExcelExporter:
    def __init__(self, project_manager, capex_manager, opex_manager, 
                 receitas_manager, config, financial_calculations):
        """
        Initialize Excel Exporter with all required managers.
        
        Args:
            project_manager: Instance of ProjectManager
            capex_manager: Instance of CapexManager
            opex_manager: Instance of OpexManager
            receitas_manager: Instance of ReceitasManager
            config: Instance of Config
            financial_calculations: Instance of FinancialCalculations
        """
        self.project_manager = project_manager
        self.capex_manager = capex_manager
        self.opex_manager = opex_manager
        self.receitas_manager = receitas_manager
        self.config = config
        self.financial_calculations = financial_calculations

    def export_project(self, filepath):
        """
        Export all project data to an Excel file.
        
        Args:
            filepath (str): Path where the Excel file will be saved
            
        Returns:
            tuple: (bool, str) - (success, message)
        """
        try:
            wb = Workbook()
            
            # Remove default sheet
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
            
            # Create and style all sheets
            self._create_summary_sheet(wb)
            self._create_capex_sheet(wb)
            self._create_opex_sheet(wb)
            self._create_receitas_sheet(wb)
            self._create_config_sheet(wb)
            self._create_results_sheet(wb)
            
            # Save workbook
            wb.save(filepath)
            return True, "Projeto exportado com sucesso"
            
        except Exception as e:
            return False, f"Erro ao exportar projeto: {str(e)}"

    def _create_summary_sheet(self, wb):
        """Create and populate the summary sheet."""
        ws = wb.create_sheet("Resumo")
        self._apply_header_style(ws, "Resumo do Projeto")
        
        # Project information
        current_project = self.project_manager.get_current_project()
        if current_project and "metadata" in current_project:
            metadata = current_project["metadata"]
            ws.append(["Nome do Projeto", metadata.get("name", "N/A")])
            ws.append(["Descrição", metadata.get("description", "N/A")])
            ws.append(["Data de Criação", metadata.get("created_at", "N/A")])
            ws.append(["Última Modificação", metadata.get("last_modified", "N/A")])
        
        ws.append([])  # Empty row
        
        # Summary of financial items
        ws.append(["Resumo Financeiro"])
        ws.append(["Categoria", "Quantidade de Itens", "Valor Total"])
        ws.append(["CapEx", len(self.capex_manager.items), self.capex_manager.total_investment])
        ws.append(["OpEx", len(self.opex_manager.items), self.opex_manager.total_annual_cost])
        ws.append(["Receitas", len(self.receitas_manager.items), self.receitas_manager.total_annual_revenue])
        
        self._auto_adjust_columns(ws)

    def _create_capex_sheet(self, wb):
        """Create and populate the CapEx sheet."""
        ws = wb.create_sheet("CapEx")
        self._apply_header_style(ws, "Capital Expenditure (CapEx)")
        
        # Headers
        headers = ["TAG", "Descrição", "Quantidade", "Valor Unitário", "Valor Total", "Mês"]
        ws.append(headers)
        self._apply_column_headers_style(ws, headers)
        
        # Data
        for item in self.capex_manager.get_all_items():
            ws.append(item.to_row())
        
        # Total
        ws.append([])
        ws.append(["Total CapEx", "", "", "", self.capex_manager.total_investment])
        
        self._auto_adjust_columns(ws)

    def _create_opex_sheet(self, wb):
        """Create and populate the OpEx sheet."""
        ws = wb.create_sheet("OpEx")
        self._apply_header_style(ws, "Operational Expenditure (OpEx)")
        
        # Headers
        headers = ["TAG", "Descrição", "Quantidade", "Valor Unitário", "Valor Total",
                  "Recorrente", "Mês Inicial", "Mês Final"]
        ws.append(headers)
        self._apply_column_headers_style(ws, headers)
        
        # Data
        for item in self.opex_manager.get_all_items():
            ws.append(item.to_row())
        
        # Total
        ws.append([])
        ws.append(["Total OpEx Anual", "", "", "", self.opex_manager.total_annual_cost])
        
        self._auto_adjust_columns(ws)

    def _create_receitas_sheet(self, wb):
        """Create and populate the Receitas sheet."""
        ws = wb.create_sheet("Receitas")
        self._apply_header_style(ws, "Receitas")
        
        # Headers
        headers = ["TAG", "Descrição", "Quantidade", "Valor Unitário", "Valor Total",
                  "Recorrente", "Mês Inicial", "Mês Final", "Taxa de Crescimento (%)"]
        ws.append(headers)
        self._apply_column_headers_style(ws, headers)
        
        # Data
        for item in self.receitas_manager.get_all_items():
            ws.append(item.to_row())
        
        # Total
        ws.append([])
        ws.append(["Total Receitas Anual", "", "", "", self.receitas_manager.total_annual_revenue])
        
        self._auto_adjust_columns(ws)

    def _create_config_sheet(self, wb):
        """Create and populate the configuration sheet."""
        ws = wb.create_sheet("Configuração")
        self._apply_header_style(ws, "Configurações do Projeto")
        
        # Tax and TMA configuration
        ws.append(["Parâmetro", "Valor"])
        self._apply_column_headers_style(ws, ["Parâmetro", "Valor"])
        
        ws.append(["TMA (%)", self.config.tma])
        ws.append(["IR (%)", self.config.ir])
        ws.append(["CSLL (%)", self.config.csll])
        ws.append(["Taxa Efetiva (%)", self.config.get_effective_tax_rate()])
        
        self._auto_adjust_columns(ws)

    def _create_results_sheet(self, wb):
        """Create and populate the results sheet."""
        ws = wb.create_sheet("Resultados")
        self._apply_header_style(ws, "Resultados da Análise")
        
        # Financial indicators
        ws.append(["Indicador", "Valor"])
        self._apply_column_headers_style(ws, ["Indicador", "Valor"])
        
        formatted_results = self.financial_calculations.format_results()
        ws.append(["TIR", formatted_results["tir"]])
        ws.append(["VPL", formatted_results["vpl"]])
        ws.append(["Payback", formatted_results["payback"]])
        ws.append(["Dívida/EBITDA", formatted_results["divida_ebitda"]])
        
        # Cash flow table
        ws.append([])
        ws.append(["Fluxo de Caixa Mensal"])
        
        headers = ["Mês", "Receitas", "OpEx", "CapEx", "EBITDA", "Impostos", "Fluxo Líquido"]
        ws.append(headers)
        self._apply_column_headers_style(ws, headers)
        
        if "cash_flows" in self.financial_calculations.results:
            for cf in self.financial_calculations.results["cash_flows"]:
                ws.append([
                    cf["month"],
                    cf["revenue"],
                    cf["opex"],
                    cf["capex"],
                    cf["ebitda"],
                    cf["taxes"],
                    cf["net_cash_flow"]
                ])
        
        self._auto_adjust_columns(ws)

    def _apply_header_style(self, ws, title):
        """Apply style to sheet header."""
        ws.append([title])
        header_cell = ws["A1"]
        header_cell.font = Font(size=14, bold=True)
        header_cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        ws.append([])  # Empty row after header

    def _apply_column_headers_style(self, ws, headers):
        """Apply style to column headers."""
        row_num = ws.max_row
        for col, _ in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
            cell.border = Border(
                bottom=Side(style='thin'),
                top=Side(style='thin'),
                left=Side(style='thin'),
                right=Side(style='thin')
            )
            cell.alignment = Alignment(horizontal="center")

    def _auto_adjust_columns(self, ws):
        """Adjust column widths based on content."""
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
