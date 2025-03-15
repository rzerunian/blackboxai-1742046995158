from datetime import datetime
from openpyxl import load_workbook
from .base_item import BaseFinancialItem

class CapExItem(BaseFinancialItem):
    def __init__(self, tag=None, description="", quantity=0, unit_price=0.0, month=1):
        """
        Initialize a CapEx item.
        
        Args:
            tag (str, optional): Unique identifier for the item
            description (str): Description of the item
            quantity (float): Quantity of the item
            unit_price (float): Unit price of the item
            month (int): Month when the investment occurs (1-12)
        """
        super().__init__(tag, description, quantity, unit_price)
        self.month = max(1, min(12, int(month)))  # Ensure month is between 1 and 12

    def to_dict(self):
        """Extend base to_dict with CapEx specific attributes."""
        data = super().to_dict()
        data["month"] = self.month
        return data

    def to_row(self):
        """Extend base to_row with CapEx specific attributes."""
        base_row = super().to_row()
        return base_row + [self.month]

    @classmethod
    def from_dict(cls, data):
        """Create a CapEx item from a dictionary."""
        item = super().from_dict(data)
        item.month = data.get("month", 1)
        return item

class CapexManager:
    def __init__(self):
        """Initialize the CapEx manager."""
        self.items = {}  # Dictionary to store items by TAG
        self.total_investment = 0.0

    def add_item(self, description, quantity, unit_price, month=1, tag=None):
        """
        Add a new CapEx item.
        
        Args:
            description (str): Description of the item
            quantity (float): Quantity of the item
            unit_price (float): Unit price of the item
            month (int): Month when the investment occurs
            tag (str, optional): Unique identifier for the item
            
        Returns:
            tuple: (bool, str) - (success, message)
        """
        try:
            item = CapExItem(tag, description, quantity, unit_price, month)
            is_valid, error_msg = item.validate()
            
            if not is_valid:
                return False, error_msg
                
            if item.tag in self.items:
                return False, "TAG já existe"
                
            self.items[item.tag] = item
            self._update_total()
            return True, "Item adicionado com sucesso"
            
        except Exception as e:
            return False, f"Erro ao adicionar item: {str(e)}"

    def update_item(self, tag, description=None, quantity=None, unit_price=None, month=None):
        """
        Update an existing CapEx item.
        
        Args:
            tag (str): TAG of the item to update
            description (str, optional): New description
            quantity (float, optional): New quantity
            unit_price (float, optional): New unit price
            month (int, optional): New month
            
        Returns:
            tuple: (bool, str) - (success, message)
        """
        try:
            if tag not in self.items:
                return False, "Item não encontrado"
                
            item = self.items[tag]
            item.update(description, quantity, unit_price)
            
            if month is not None:
                item.month = max(1, min(12, int(month)))
                
            is_valid, error_msg = item.validate()
            if not is_valid:
                return False, error_msg
                
            self._update_total()
            return True, "Item atualizado com sucesso"
            
        except Exception as e:
            return False, f"Erro ao atualizar item: {str(e)}"

    def delete_item(self, tag):
        """
        Delete a CapEx item.
        
        Args:
            tag (str): TAG of the item to delete
            
        Returns:
            tuple: (bool, str) - (success, message)
        """
        try:
            if tag not in self.items:
                return False, "Item não encontrado"
                
            del self.items[tag]
            self._update_total()
            return True, "Item removido com sucesso"
            
        except Exception as e:
            return False, f"Erro ao remover item: {str(e)}"

    def get_item(self, tag):
        """
        Get a specific CapEx item.
        
        Args:
            tag (str): TAG of the item
            
        Returns:
            CapExItem: The item if found, None otherwise
        """
        return self.items.get(tag)

    def get_all_items(self):
        """
        Get all CapEx items.
        
        Returns:
            list: List of all CapEx items
        """
        return list(self.items.values())

    def get_monthly_investment(self, month):
        """
        Get total investment for a specific month.
        
        Args:
            month (int): Month number (1-12)
            
        Returns:
            float: Total investment for the month
        """
        return sum(item.total_value for item in self.items.values() if item.month == month)

    def get_monthly_investments(self):
        """
        Get investments for all months.
        
        Returns:
            list: List of 12 values representing monthly investments
        """
        monthly = [0.0] * 12
        for item in self.items.values():
            monthly[item.month - 1] += item.total_value
        return monthly

    def import_from_excel(self, filepath):
        """
        Import CapEx items from an Excel file.
        
        Args:
            filepath (str): Path to the Excel file
            
        Returns:
            tuple: (bool, str) - (success, message)
        """
        try:
            wb = load_workbook(filepath, read_only=True)
            ws = wb.active
            
            # Skip header row
            rows = list(ws.rows)[1:]
            
            success_count = 0
            error_count = 0
            
            for row in rows:
                try:
                    # Assuming Excel structure: TAG, Description, Quantity, Unit Price, Month
                    tag = row[0].value
                    description = row[1].value
                    quantity = float(row[2].value or 0)
                    unit_price = float(row[3].value or 0)
                    month = int(row[4].value or 1)
                    
                    success, _ = self.add_item(description, quantity, unit_price, month, tag)
                    if success:
                        success_count += 1
                    else:
                        error_count += 1
                        
                except Exception:
                    error_count += 1
                    
            return True, f"Importação concluída. Sucesso: {success_count}, Erros: {error_count}"
            
        except Exception as e:
            return False, f"Erro ao importar arquivo: {str(e)}"

    def export_to_excel(self, filepath):
        """
        Export CapEx items to an Excel file.
        
        Args:
            filepath (str): Path to save the Excel file
            
        Returns:
            tuple: (bool, str) - (success, message)
        """
        try:
            from openpyxl import Workbook
            
            wb = Workbook()
            ws = wb.active
            
            # Write headers
            headers = ["TAG", "Descrição", "Quantidade", "Valor Unitário", "Valor Total", "Mês"]
            ws.append(headers)
            
            # Write items
            for item in self.items.values():
                ws.append(item.to_row())
            
            wb.save(filepath)
            return True, "Exportação concluída com sucesso"
            
        except Exception as e:
            return False, f"Erro ao exportar arquivo: {str(e)}"

    def _update_total(self):
        """Update the total investment value."""
        self.total_investment = sum(item.total_value for item in self.items.values())

    def clear(self):
        """Clear all items."""
        self.items.clear()
        self.total_investment = 0.0
