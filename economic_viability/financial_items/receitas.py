from datetime import datetime
from openpyxl import load_workbook
from .base_item import BaseFinancialItem

class ReceitaItem(BaseFinancialItem):
    def __init__(self, tag=None, description="", quantity=0, unit_price=0.0, 
                 recurrent=True, start_month=1, end_month=12, growth_rate=0.0):
        """
        Initialize a Receita (Revenue) item.
        
        Args:
            tag (str, optional): Unique identifier for the item
            description (str): Description of the item
            quantity (float): Quantity of the item
            unit_price (float): Unit price of the item
            recurrent (bool): Whether the revenue is recurrent monthly
            start_month (int): Month when the revenue starts (1-12)
            end_month (int): Month when the revenue ends (1-12)
            growth_rate (float): Monthly growth rate in percentage (0-100)
        """
        super().__init__(tag, description, quantity, unit_price)
        self.recurrent = recurrent
        self.start_month = max(1, min(12, int(start_month)))
        self.end_month = max(self.start_month, min(12, int(end_month)))
        self.growth_rate = max(0.0, min(100.0, float(growth_rate)))

    def to_dict(self):
        """Extend base to_dict with Receita specific attributes."""
        data = super().to_dict()
        data.update({
            "recurrent": self.recurrent,
            "start_month": self.start_month,
            "end_month": self.end_month,
            "growth_rate": self.growth_rate
        })
        return data

    def to_row(self):
        """Extend base to_row with Receita specific attributes."""
        base_row = super().to_row()
        return base_row + [self.recurrent, self.start_month, self.end_month, self.growth_rate]

    @classmethod
    def from_dict(cls, data):
        """Create a Receita item from a dictionary."""
        item = super().from_dict(data)
        item.recurrent = data.get("recurrent", True)
        item.start_month = data.get("start_month", 1)
        item.end_month = data.get("end_month", 12)
        item.growth_rate = data.get("growth_rate", 0.0)
        return item

    def get_monthly_revenue(self, month):
        """
        Calculate the revenue for a specific month, including growth.
        
        Args:
            month (int): Month number (1-12)
            
        Returns:
            float: Revenue for the specified month
        """
        if self.start_month <= month <= self.end_month:
            months_passed = month - self.start_month
            growth_factor = (1 + self.growth_rate/100) ** months_passed
            return self.total_value * growth_factor
        return 0.0

class ReceitasManager:
    def __init__(self):
        """Initialize the Receitas manager."""
        self.items = {}  # Dictionary to store items by TAG
        self.total_annual_revenue = 0.0

    def add_item(self, description, quantity, unit_price, recurrent=True,
                 start_month=1, end_month=12, growth_rate=0.0, tag=None):
        """
        Add a new Receita item.
        
        Args:
            description (str): Description of the item
            quantity (float): Quantity of the item
            unit_price (float): Unit price of the item
            recurrent (bool): Whether the revenue is recurrent monthly
            start_month (int): Month when the revenue starts
            end_month (int): Month when the revenue ends
            growth_rate (float): Monthly growth rate in percentage
            tag (str, optional): Unique identifier for the item
            
        Returns:
            tuple: (bool, str) - (success, message)
        """
        try:
            item = ReceitaItem(tag, description, quantity, unit_price,
                             recurrent, start_month, end_month, growth_rate)
            is_valid, error_msg = item.validate()
            
            if not is_valid:
                return False, error_msg
                
            if item.tag in self.items:
                return False, "TAG já existe"
                
            self.items[item.tag] = item
            self._update_total()
            return True, "Item adicionado com sucesso"
            
        except Exception as e:
            return False, f"Erro ao adicionar item: {str(e)}"

    def update_item(self, tag, description=None, quantity=None, unit_price=None,
                   recurrent=None, start_month=None, end_month=None, growth_rate=None):
        """
        Update an existing Receita item.
        
        Args:
            tag (str): TAG of the item to update
            description (str, optional): New description
            quantity (float, optional): New quantity
            unit_price (float, optional): New unit price
            recurrent (bool, optional): New recurrent status
            start_month (int, optional): New start month
            end_month (int, optional): New end month
            growth_rate (float, optional): New growth rate
            
        Returns:
            tuple: (bool, str) - (success, message)
        """
        try:
            if tag not in self.items:
                return False, "Item não encontrado"
                
            item = self.items[tag]
            item.update(description, quantity, unit_price)
            
            if recurrent is not None:
                item.recurrent = recurrent
            if start_month is not None:
                item.start_month = max(1, min(12, int(start_month)))
            if end_month is not None:
                item.end_month = max(item.start_month, min(12, int(end_month)))
            if growth_rate is not None:
                item.growth_rate = max(0.0, min(100.0, float(growth_rate)))
                
            is_valid, error_msg = item.validate()
            if not is_valid:
                return False, error_msg
                
            self._update_total()
            return True, "Item atualizado com sucesso"
            
        except Exception as e:
            return False, f"Erro ao atualizar item: {str(e)}"

    def delete_item(self, tag):
        """
        Delete a Receita item.
        
        Args:
            tag (str): TAG of the item to delete
            
        Returns:
            tuple: (bool, str) - (success, message)
        """
        try:
            if tag not in self.items:
                return False, "Item não encontrado"
                
            del self.items[tag]
            self._update_total()
            return True, "Item removido com sucesso"
            
        except Exception as e:
            return False, f"Erro ao remover item: {str(e)}"

    def get_item(self, tag):
        """
        Get a specific Receita item.
        
        Args:
            tag (str): TAG of the item
            
        Returns:
            ReceitaItem: The item if found, None otherwise
        """
        return self.items.get(tag)

    def get_all_items(self):
        """
        Get all Receita items.
        
        Returns:
            list: List of all Receita items
        """
        return list(self.items.values())

    def get_monthly_revenue(self, month):
        """
        Get total revenue for a specific month.
        
        Args:
            month (int): Month number (1-12)
            
        Returns:
            float: Total revenue for the month
        """
        return sum(item.get_monthly_revenue(month) for item in self.items.values())

    def get_monthly_revenues(self):
        """
        Get revenues for all months.
        
        Returns:
            list: List of 12 values representing monthly revenues
        """
        return [self.get_monthly_revenue(month) for month in range(1, 13)]

    def import_from_excel(self, filepath):
        """
        Import Receita items from an Excel file.
        
        Args:
            filepath (str): Path to the Excel file
            
        Returns:
            tuple: (bool, str) - (success, message)
        """
        try:
            wb = load_workbook(filepath, read_only=True)
            ws = wb.active
            
            # Skip header row
            rows = list(ws.rows)[1:]
            
            success_count = 0
            error_count = 0
            
            for row in rows:
                try:
                    # Assuming Excel structure: TAG, Description, Quantity, Unit Price, Recurrent, Start Month, End Month, Growth Rate
                    tag = row[0].value
                    description = row[1].value
                    quantity = float(row[2].value or 0)
                    unit_price = float(row[3].value or 0)
                    recurrent = bool(row[4].value)
                    start_month = int(row[5].value or 1)
                    end_month = int(row[6].value or 12)
                    growth_rate = float(row[7].value or 0)
                    
                    success, _ = self.add_item(
                        description, quantity, unit_price, recurrent,
                        start_month, end_month, growth_rate, tag
                    )
                    if success:
                        success_count += 1
                    else:
                        error_count += 1
                        
                except Exception:
                    error_count += 1
                    
            return True, f"Importação concluída. Sucesso: {success_count}, Erros: {error_count}"
            
        except Exception as e:
            return False, f"Erro ao importar arquivo: {str(e)}"

    def export_to_excel(self, filepath):
        """
        Export Receita items to an Excel file.
        
        Args:
            filepath (str): Path to save the Excel file
            
        Returns:
            tuple: (bool, str) - (success, message)
        """
        try:
            from openpyxl import Workbook
            
            wb = Workbook()
            ws = wb.active
            
            # Write headers
            headers = ["TAG", "Descrição", "Quantidade", "Valor Unitário", "Valor Total",
                      "Recorrente", "Mês Inicial", "Mês Final", "Taxa de Crescimento (%)"]
            ws.append(headers)
            
            # Write items
            for item in self.items.values():
                ws.append(item.to_row())
            
            wb.save(filepath)
            return True, "Exportação concluída com sucesso"
            
        except Exception as e:
            return False, f"Erro ao exportar arquivo: {str(e)}"

    def _update_total(self):
        """Update the total annual revenue."""
        monthly_revenues = self.get_monthly_revenues()
        self.total_annual_revenue = sum(monthly_revenues)

    def clear(self):
        """Clear all items."""
        self.items.clear()
        self.total_annual_revenue = 0.0
